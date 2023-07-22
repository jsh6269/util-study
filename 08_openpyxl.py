from openpyxl import Workbook, load_workbook

wb = Workbook()
print(wb.sheetnames)

# 시트 이름 바꾸기
wb['Sheet'].title = 'Exercise'

# 시트 추가, 삭제
wb.create_sheet('Ex1')
wb.remove(wb['Exercise'])
print(wb.sheetnames)

# 행 추가
wb['Ex1'].append([1, 2, 3])
wb['Ex1'].append([4, 5, 6])
wb['Ex1'].append([7, 8, 9])

wb.save('sample/output.xlsx')

# 엑셀 불러오기 (data_only를 true 설정시 수식이 아닌 값으로 불러옴)
load_wb = load_workbook('sample/Stock.xlsx', data_only=True)
stock = load_wb['Stock']
for row in stock.rows:
    for cell in row:
        print(cell.value, end=' ')
    print()
